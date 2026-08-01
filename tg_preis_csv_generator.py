# -*- coding: utf-8 -*-
"""Neue TG-Preise (VC000153) als CSV fuer den CALKU-Preisimport aufbereiten.

Der CSV-Import in CALKU matcht ueber den REZEPT-Zutatnamen und erwartet
den Preis je kg/Liter. Diese Aufbereitung uebersetzt deshalb:

  Rezept-Zutat  ->  CALKU-Preislisteneintrag (Namensaehnlichkeit >= 0,8)
                ->  TG-Artikel (gleiche Artikelnummer)
                ->  neuer kg-Preis = alter kg-Preis x (TG neu / Gebinde alt)

Die Skalierung erhaelt die bewaehrte Gebinde-Interpretation des alten
Eintrags. Bewusst AUSGESCHLOSSEN (stehen im Abgleich-Report, Blatt 1 rot):
  - TG-Einheit ist nicht ST (Colli o. ae. - Gebinde unklar)
  - Abweichung > 60 Prozent (Verdacht Gebindewechsel)
  - die Presswaren-Zutaten (Frischer Apfelsaft/Orangensaft, Karottensaft) -
    deren Netto-Kilopreise pflegt Mark gerade selbst mit den Auspressquoten

Ausgabe: Downloads/calku_tg_preise_<datum>.csv (Semikolon, Dezimalkomma)
"""
import csv
import json
import sys
from datetime import date
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, r"C:\Users\Media\Apps\08_Tech_und_Tools\immergruen-bigquery-pipeline")

from openpyxl import load_workbook

from igorder.normalisierung import bester_treffer

HIER = Path(__file__).parent
TG_DATEI = Path.home() / "Downloads" / "Artikelliste VC000153.xlsx"
ZIEL = Path.home() / "Downloads" / f"calku_tg_preise_{date.today().isoformat()}.csv"

AUSGESCHLOSSENE_ZUTATEN = {"frischer apfelsaft", "frischer orangensaft",
                           "karottensaft"}
MAX_ABWEICHUNG_PROZENT = 60


def norm_nr(wert):
    s = str(wert or "").strip()
    return s[:-2] if s.endswith(".0") else (s or None)


# TG-Liste
wb = load_workbook(TG_DATEI, data_only=True, read_only=True)
tg = {}
for zeile in wb.active.iter_rows(min_row=2, values_only=True):
    nr, text, me, preis = (list(zeile) + [None] * 4)[:4]
    nr = norm_nr(nr)
    if nr and preis is not None:
        tg[nr] = {"text": str(text or ""), "me": str(me or ""),
                  "preis": float(preis)}

# CALKU-Preisliste (Basis, wie die App sie baut)
calku = {}
for datei in ["rezeptdatenbank.json", "smoothies_v3.json", "juices_v3.json",
              "refresher_v1.json"]:
    inhalt = json.loads((HIER / "src" / "data" / datei).read_text(encoding="utf-8"))
    for e in inhalt.get("price_list", []):
        if e.get("ingredient_name"):
            calku[e["ingredient_name"].lower()] = e

# Rezept-Zutatennamen (das ist der Match-Schluessel des CSV-Imports)
stand = json.loads((HIER / "src" / "data" / "stand_2026-06-05.json")
                   .read_text(encoding="utf-8"))
zutaten = sorted({z["name"] for p in stand["produkte"]
                  for z in p.get("zutaten", [])})

liste_namen = list(calku.keys())
zeilen, ausgelassen = [], []
for zutat in zutaten:
    if zutat.lower() in AUSGESCHLOSSENE_ZUTATEN:
        ausgelassen.append((zutat, "Presswaren-Zutat, Pflege bei Mark"))
        continue
    treffer_name, score = bester_treffer(zutat, liste_namen, schwelle=0.8)
    if not treffer_name:
        continue
    eintrag = calku[treffer_name]
    nr = norm_nr(eintrag.get("article_number"))
    if not nr or nr not in tg:
        continue
    neu = tg[nr]
    alt_gebinde = eintrag.get("package_price")
    alt_pro_g = eintrag.get("price_per_gram_ml")
    if not alt_gebinde or not alt_pro_g:
        continue
    if neu["me"] != "ST":
        ausgelassen.append((zutat, f"TG-Einheit {neu['me']} (Gebinde unklar)"))
        continue
    delta_pct = (neu["preis"] - alt_gebinde) / alt_gebinde * 100
    if abs(delta_pct) > MAX_ABWEICHUNG_PROZENT:
        ausgelassen.append((zutat, f"Abweichung {delta_pct:+.0f} % (Gebindewechsel?)"))
        continue
    neuer_kg = alt_pro_g * 1000 * (neu["preis"] / alt_gebinde)
    zeilen.append({
        "Artikelname": zutat,
        "Preis pro kg (EUR)": f"{neuer_kg:.4f}".replace(".", ","),
        "Einheit": "kg",
        "Artikelnummer": nr,
        "TG-Bezeichnung": neu["text"],
        "alt pro kg": f"{alt_pro_g * 1000:.4f}".replace(".", ","),
        "Aenderung %": f"{delta_pct:+.1f}".replace(".", ","),
    })

with open(ZIEL, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(zeilen[0].keys()), delimiter=";")
    w.writeheader()
    w.writerows(zeilen)

geaendert = [z for z in zeilen if z["Aenderung %"] not in ("+0,0", "-0,0")]
print(f"CSV: {ZIEL}")
print(f"Zeilen gesamt: {len(zeilen)}  |  davon mit Preisaenderung: {len(geaendert)}")
print("\nGroesste Aenderungen im CSV:")
for z in sorted(geaendert,
                key=lambda x: -abs(float(x["Aenderung %"].replace(",", ".").replace("+", ""))))[:12]:
    print(f"  {z['Artikelname']:30s} {z['alt pro kg']:>10s} -> "
          f"{z['Preis pro kg (EUR)']:>10s} EUR/kg  ({z['Aenderung %']} %)")
print(f"\nBewusst ausgelassen ({len(ausgelassen)}):")
for name, grund in ausgelassen:
    print(f"  {name}: {grund}")
