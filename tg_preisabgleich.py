# -*- coding: utf-8 -*-
"""Abgleich: neue TG-Artikelliste (VC000153) gegen die CALKU-Preisliste.

Rekonstruiert die effektive CALKU-Preisliste wie die App (rezeptdatenbank
+ Overrides aus smoothies_v3/juices_v3/refresher_v1; manuell in der App
angelegte Artikel liegen nur in Supabase und fehlen hier - kommen ueber
Marks Export nach).

Ergebnis-Excel mit drei Blaettern:
  1. Preisabweichungen  - gleiche Artikelnummer, anderer Preis
  2. Dubletten          - mehrere CALKU-Eintraege je Artikelnummer/Name,
                          mit Empfehlung behalten/loeschen
  3. Ohne TG-Treffer    - CALKU-Eintraege ohne Nummer in der TG-Liste,
                          mit bestem Namensvorschlag aus der TG-Liste

Ausfuehren:  py -3 tg_preisabgleich.py
"""
import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, r"C:\Users\Media\Apps\08_Tech_und_Tools\immergruen-bigquery-pipeline")

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from igorder.normalisierung import normalisiere, bester_treffer, groesse_ml

HIER = Path(__file__).parent
TG_DATEI = Path.home() / "Downloads" / "Artikelliste VC000153.xlsx"
ZIEL = Path.home() / "Downloads" / f"tg_preisabgleich_{date.today().isoformat()}.json".replace(".json", ".xlsx")

GRUEN = "FF084B32"
ROT = "FFFDE7E9"


def norm_nr(wert):
    if wert is None:
        return None
    s = str(wert).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s or None


# ---------------------------------------------------------------- TG-Liste
wb = load_workbook(TG_DATEI, data_only=True, read_only=True)
ws = wb.active
tg = {}
for zeile in ws.iter_rows(min_row=2, values_only=True):
    nr, text, me, preis = (list(zeile) + [None] * 4)[:4]
    nr = norm_nr(nr)
    if nr and preis is not None:
        tg[nr] = {"text": str(text or ""), "me": str(me or ""), "preis": float(preis)}
print(f"TG-Liste: {len(tg)} Artikel")

# ------------------------------------------------- CALKU-Preisliste (Basis)
quellen = ["rezeptdatenbank.json", "smoothies_v3.json", "juices_v3.json",
           "refresher_v1.json"]
calku = {}
for datei in quellen:
    inhalt = json.loads((HIER / "src" / "data" / datei).read_text(encoding="utf-8"))
    for eintrag in inhalt.get("price_list", []):
        name = eintrag.get("ingredient_name")
        if name:
            calku[name.lower()] = {**eintrag, "_quelle": datei}
print(f"CALKU-Preisliste (Basis): {len(calku)} Eintraege")

eintraege = []
for schluessel, e in sorted(calku.items()):
    eintraege.append({
        "name": e.get("ingredient_name", ""),
        "nr": norm_nr(e.get("article_number")),
        "einheit": e.get("unit") or "",
        "menge": e.get("package_size"),
        "preis": e.get("package_price"),
        "pro_g": e.get("price_per_gram_ml"),
        "geprueft": str(e.get("date_last_checked") or "")[:10],
        "quelle": e.get("_quelle", ""),
    })

# ------------------------------------------------------------- Auswertung
abweichungen, ohne_treffer = [], []
je_nr = defaultdict(list)
for e in eintraege:
    if e["nr"] and e["nr"] in tg:
        je_nr[e["nr"]].append(e)
        neu = tg[e["nr"]]["preis"]
        alt = e["preis"]
        if alt is not None and abs(neu - alt) > 0.005:
            abweichungen.append({**e, "neu": neu,
                                 "delta": neu - alt,
                                 "delta_pct": (neu - alt) / alt * 100 if alt else 0,
                                 "tg_text": tg[e["nr"]]["text"]})
    else:
        kandidaten = list(tg.values())
        namen = [k["text"] for k in kandidaten]
        treffer, score = bester_treffer(e["name"], namen, schwelle=0.55)
        vorschlag = ""
        v_nr, v_preis = "", None
        if treffer:
            for nr, k in tg.items():
                if k["text"] == treffer:
                    vorschlag, v_nr, v_preis = treffer, nr, k["preis"]
                    break
        ohne_treffer.append({**e, "vorschlag": vorschlag, "v_nr": v_nr,
                             "v_preis": v_preis, "score": round(score, 2)})

# Dubletten: gleiche TG-Nummer mehrfach ODER gleicher normalisierter Name
gruppen = []
for nr, liste in je_nr.items():
    if len(liste) > 1:
        gruppen.append(("Artikelnummer " + nr, liste))
def ist_groessenfamilie(liste):
    """Becher 0,4/0,5/0,6 l sind Varianten, keine Dubletten."""
    groessen = {groesse_ml(e["name"]) for e in liste}
    groessen.discard(None)
    return len(groessen) > 1


je_name = defaultdict(list)
for e in eintraege:
    je_name[normalisiere(e["name"])].append(e)
for schluessel, liste in je_name.items():
    if len(liste) > 1 and not ist_groessenfamilie(liste) \
            and not any(schluessel in g[0] for g in gruppen):
        gruppen.append(("Name „" + liste[0]["name"] + "…", liste))

# zusaetzlich: sehr aehnliche Namen (Banane/Bananen, Mixsalat/Salatmix ...)
namensliste = [e["name"] for e in eintraege]
gesehen = set()
for e in eintraege:
    if e["name"] in gesehen:
        continue
    andere = [n for n in namensliste if n != e["name"]]
    treffer, score = bester_treffer(e["name"], andere, schwelle=0.86)
    if treffer and (treffer, e["name"]) not in gesehen:
        paar = sorted([e["name"], treffer])
        schluessel = tuple(paar)
        if schluessel not in gesehen:
            gesehen.add(schluessel)
            gesehen.add(e["name"]); gesehen.add(treffer)
            mitglieder = [x for x in eintraege if x["name"] in paar]
            if ist_groessenfamilie(mitglieder):
                continue
            if len({m["name"] for m in mitglieder}) > 1:
                bereits = any(set(m["name"] for m in mitglieder)
                              <= set(x["name"] for x in g[1]) for g in gruppen)
                if not bereits:
                    gruppen.append((f"Ähnliche Namen ({int(score*100)} %)", mitglieder))

# -------------------------------------------------------------- Excel-Report
mappe = Workbook()


def kopf(blatt, spalten):
    blatt.append([s[0] for s in spalten])
    for i, (titel, breite) in enumerate(spalten, start=1):
        z = blatt.cell(row=1, column=i)
        z.font = Font(bold=True, color="FFFFFFFF")
        z.fill = PatternFill("solid", fgColor=GRUEN)
        z.alignment = Alignment(vertical="center")
        blatt.column_dimensions[z.column_letter].width = breite
    blatt.freeze_panes = "A2"


b1 = mappe.active
b1.title = "Preisabweichungen"
kopf(b1, [("Zutat (CALKU)", 32), ("Artikel-Nr.", 12), ("TG-Bezeichnung", 30),
          ("Preis CALKU", 12), ("Preis TG neu", 12), ("TG-Einheit", 11),
          ("Δ €", 10), ("Δ %", 10), ("zuletzt geprüft", 14)])
for a in sorted(abweichungen, key=lambda x: -abs(x["delta_pct"])):
    b1.append([a["name"], a["nr"], a["tg_text"], a["preis"], a["neu"],
               tg[a["nr"]]["me"], round(a["delta"], 2),
               round(a["delta_pct"], 1), a["geprueft"]])
    # Colli-Einheit oder Extremsprung: vermutlich Gebindewechsel, kein Preis
    if tg[a["nr"]]["me"] != "ST" or abs(a["delta_pct"]) > 100:
        b1.cell(row=b1.max_row, column=6).fill = PatternFill("solid", fgColor=ROT)

b2 = mappe.create_sheet("Dubletten")
kopf(b2, [("Gruppe", 26), ("Zutat (CALKU)", 32), ("Artikel-Nr.", 12),
          ("Preis", 10), ("Quelle", 22), ("Empfehlung", 34)])
for titel, liste in gruppen:
    # behalten: Eintrag mit TG-Treffer und juengstem Pruefdatum
    beste = sorted(liste, key=lambda e: ((e["nr"] in tg), e["geprueft"]),
                   reverse=True)[0]
    for e in liste:
        empfehlung = ("BEHALTEN (TG-Treffer, aktuellster Stand)"
                      if e is beste else "löschen / Rezepte auf den anderen umstellen")
        zeile = [titel, e["name"], e["nr"] or "—", e["preis"], e["quelle"],
                 empfehlung]
        b2.append(zeile)
        if e is not beste:
            b2.cell(row=b2.max_row, column=6).fill = PatternFill("solid", fgColor=ROT)

b3 = mappe.create_sheet("Ohne TG-Treffer")
kopf(b3, [("Zutat (CALKU)", 32), ("Artikel-Nr. (alt)", 14), ("Preis", 10),
          ("Vorschlag aus TG-Liste", 30), ("TG-Nr.", 12), ("TG-Preis", 10),
          ("Güte", 8)])
for e in sorted(ohne_treffer, key=lambda x: -(x["score"] or 0)):
    b3.append([e["name"], e["nr"] or "—", e["preis"], e["vorschlag"],
               e["v_nr"], e["v_preis"], e["score"]])

mappe.save(ZIEL)

print(f"\nPreisabweichungen : {len(abweichungen)}")
print(f"Dubletten-Gruppen : {len(gruppen)}")
print(f"Ohne TG-Treffer   : {len(ohne_treffer)}")
print(f"\nReport: {ZIEL}")
print("\nGroesste Abweichungen:")
for a in sorted(abweichungen, key=lambda x: -abs(x["delta_pct"]))[:10]:
    print(f"  {a['name']:32s} {a['preis']:8.2f} -> {a['neu']:8.2f}  ({a['delta_pct']:+.0f} %)")
print("\nDubletten:")
for titel, liste in gruppen[:12]:
    print(f"  {titel}: " + " | ".join(e['name'] for e in liste))
